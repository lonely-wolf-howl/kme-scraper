import psutil
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')
# 현재 사용자의 MAC 주소를 추출하는 함수입니다.
def get_mac_address():
    interfaces = psutil.net_if_addrs()
    for interface in interfaces.values():
        for addr in interface:
            if addr.family == psutil.AF_LINK:
                return addr.address
    return None

# 추출된 사용자의 MAC 주소와 초기 설정된 MAC 값을 비교합니다.
user_mac_address = get_mac_address()
print('this user => MAC : ' + user_mac_address)

default_mac_address = "40-B0-76-42-8F-7D" # <--- 사용자를 통해 전달받고 설정해야 하는 값입니다!
print('default => MAC : ' + default_mac_address)

# MAC 주소가 일치하는지 확인하여, 실행 여부를 결정합니다.
if user_mac_address == default_mac_address:
    # 현재 날짜와 초기 설정된 날짜를 비교하여, 사용 가능 기간인지 확인합니다.
    default_set_date = datetime.datetime.strptime("2023-05-11", "%Y-%m-%d").date()
    current_date = datetime.datetime.now().date()
    expiration_date = default_set_date + datetime.timedelta(days=30) # <--- 초기 설정된 날짜로부터 30일 후

    print('')
    print('default set date : ' + str(default_set_date))
    print('current date : ' + str(current_date))
    print('expiration date : ' + str(expiration_date))
    print('')

    if current_date <= expiration_date: # <<<============================================================
        print('정상적으로 실행되었습니다!')
        print('')
        import customtkinter as ctk # 'customtkinter'를 가져옵니다.
        import openpyxl # Excel 자료의 읽기, 쓰기, 수정 기능을 제공합니다.
        import tkinter as tk # GUI 기능을 제공합니다.
        from tkinter import ttk, messagebox
        import re # pattern을 사용하여 문자열을 조작하고, 그에 기반한 작업을 수행할 수 있습니다.
        from PIL import Image # 사진 객체를 만들기 위해 불러옵니다.
        from PIL import ImageTk # 'tkinter canvas'에서 사진 객체에 대한 다양한 작업을 제공합니다.

        import pandas as pd # 자료 조작과 분석을 위해 가져옵니다. 행렬을 손쉽게 생성할 수 있습니다.

        # 실행 시 자동으로 'googletrans'를 최신화 합니다.
        import subprocess
        import pkg_resources

        installed_packages = [package.project_name for package in pkg_resources.working_set]
        if 'googletrans' in installed_packages:
            package_name = 'googletrans'
            current_version = next((package.version for package in pkg_resources.working_set if package.project_name == package_name), None)
            if current_version and current_version != '4.0.0-rc1':
                subprocess.check_call(['pip', 'install', '--upgrade', 'googletrans==4.0.0-rc1'], 
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        # class => top-level window
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        toplevel_window = None  # 전역 변수로 선언합니다.

        class ToplevelWindow(ctk.CTkToplevel):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.title("경고")
                self.geometry("300x150")
                self.grid_rowconfigure((0, 1), weight=1)
                self.grid_columnconfigure((0, 1), weight=1)

                self.label = ctk.CTkLabel(self, text="정말로 종료하시겠습니까?", font=font_style)
                self.label.grid(row=0, column=0, columnspan=2, padx=20, pady=(20,10), sticky="ew")

                self.button = ctk.CTkButton(self, text="예", width=100, font=font_style, 
                command=close_all)
                self.button.grid(row=1, column=0, padx=(20,10), pady=(10,20), sticky="ew")

                self.button = ctk.CTkButton(self, text="아니오", width=100, font=font_style, 
                command=close_toplevel)
                self.button.grid(row=1, column=1, padx=(10,20), pady=(10,20), sticky="ew")

        def open_toplevel(): # <--- 창 종료를 눌렀을 때, 실행되는 함수입니다.
            global toplevel_window

            if toplevel_window is None or not toplevel_window.winfo_exists():
                toplevel_window = ToplevelWindow(root)
                toplevel_window.attributes('-topmost', True) # 추가 창을 항상 맨 위에 표시합니다.
            else:
                toplevel_window.lift() # 이미 존재하는 창을 맨 위로 올립니다.
                toplevel_window.focus() # 추가 창에 초점을 맞춥니다.

        def close_toplevel(): # <--- 추가 창을 종료하는 함수입니다.
            if toplevel_window is not None and toplevel_window.winfo_exists():
                toplevel_window.destroy() # 추가 창을 종료합니다.

        def close_all(): # <--- 추가 창, 기본 창을 모두 종료하는 함수입니다.
            if toplevel_window is not None and toplevel_window.winfo_exists():
                toplevel_window.destroy() # 추가 창을 종료합니다.
                root.destroy() # 기본 창을 종료합니다.

        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        # function
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        amazon_urls = [] # 입력되는 'amazon_url'을 받기 위한 빈 배열을 선언합니다.
        iherb_urls = [] # 입력되는 'iherb_url'을 받기 위한 빈 배열을 선언합니다.

        thumbnail_color = 'white' # thumbnail 색상 초기값
        # ================================================================================
        '''
        출력창에 자료를 행렬로 보여주는 함수
        '''
        def load_data():
            amazon_iherb_value = amazon_iherb_option_var.get()

            if amazon_iherb_value == "amazon":
                path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
                workbook = pd.read_excel(path, sheet_name='amazon')

                pd.set_option('display.max_rows', None)
                pd.set_option('display.max_columns', None)

                pd.set_option('display.width', None)

                print(workbook)

            else:
                path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
                workbook = pd.read_excel(path, sheet_name='iherb')

                pd.set_option('display.max_rows', None)
                pd.set_option('display.max_columns', None)

                pd.set_option('display.width', None)

                print(workbook)
        # ================================================================================
        '''
        thumbnail 테두리에 들어갈 색상 값을 가져오는 함수
        '''
        def color_segmented_button_callback(value):
            global thumbnail_color
            thumbnail_color = value.strip()
        # ================================================================================
        '''
        제품 주소가 중복되는지 확인하고, 목록에 추가하는 함수
        '''
        def check_duplicates_and_add_URL():
            amazon_iherb_value = amazon_iherb_option_var.get()

            if amazon_iherb_value == "amazon":
                amazon_url = url_entry.get().strip()
                wb = openpyxl.load_workbook("./product_list.xlsx")
                sheet = wb['amazon']
                product_list_amazon_urls = []
                for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
                    product_list_amazon_urls.append(row[0].value)

                log_textbox.delete("0.0", ctk.END)
                if amazon_url in product_list_amazon_urls:
                    answer = "[경고] 중복되는 제품 주소가 존재합니다!"
                    log_textbox.insert("0.0", answer)
                else:
                    answer = "제품 주소가 추가되었습니다."
                    log_textbox.insert("0.0", answer)

                    amazon_urls.append(amazon_url) # <--- 입력된 주소가 배열(amazon_urls)에 저장됩니다!
                    # print(amazon_urls)

                    # 제품 주소 = frame
                    url_frame = ctk.CTkFrame(url_scrollable_frame)
                    url_frame.pack(fill="x", pady=(5,0))

                    # 삭제 = button
                    delete_button = ctk.CTkButton(url_frame, text="삭제", width=50, fg_color="#CC3D3D", hover_color="#960707", font=font_style, 
                                        command=lambda frame=url_frame, url=amazon_url: 
                                        (amazon_urls.remove(url), log_textbox.delete("0.0", ctk.END), log_textbox.insert("0.0", "삭제 완료!"), frame.destroy()))
                    delete_button.pack(side="left", padx=5, pady=5)

                    # 제품 번호(asin_code) = button
                    asin_code = re.search(r"dp\/([A-Z0-9]{10})\/", amazon_url).group(1)
                    asin_code_button = ctk.CTkButton(url_frame, text=f"{asin_code}", width=50, font=font_style)
                    asin_code_button.pack(side="left", pady=5)

                    # 제품 주소 = label
                    label = ctk.CTkLabel(url_frame, text=amazon_url, font=font_style)
                    label.pack(side="left", padx=5, pady=5, anchor="center")

                    return amazon_urls # <--- 제품 주소들이 누적된 배열을 반환합니다!

            else:
                iherb_url = url_entry.get().strip()
                wb = openpyxl.load_workbook("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx")
                sheet = wb['iherb']
                product_list_iherb_urls = []
                for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
                    product_list_iherb_urls.append(row[0].value)

                log_textbox.delete("0.0", ctk.END)
                if iherb_url in product_list_iherb_urls:
                    answer = "[경고] 중복되는 제품 주소가 존재합니다!"
                    log_textbox.insert("0.0", answer)
                else:
                    answer = "제품 주소가 추가되었습니다."
                    log_textbox.insert("0.0", answer)

                    iherb_urls.append(iherb_url) # <--- 입력된 주소가 배열(amazon_urls)에 저장됩니다!
                    # print(iherb_urls)

                    # 제품 주소 = frame
                    url_frame = ctk.CTkFrame(url_scrollable_frame)
                    url_frame.pack(fill="x", pady=(5,0))

                    # 삭제 = button
                    delete_button = ctk.CTkButton(url_frame, text="삭제", width=50, fg_color="#CC3D3D", hover_color="#960707", font=font_style, 
                                        command=lambda frame=url_frame, url=iherb_url: 
                                        (iherb_urls.remove(url), log_textbox.delete("0.0", ctk.END), log_textbox.insert("0.0", "삭제 완료!"), frame.destroy()))
                    delete_button.pack(side="left", padx=5, pady=5)

                    # 제품 번호(product_id) = button
                    product_id = iherb_url.rsplit('/', 1)[-1].replace("?rec=home", "")
                    # product_id = re.search(r"/(\d+)$", iherb_url).group(1)
                    product_id_button = ctk.CTkButton(url_frame, text=f"{product_id}", width=50, font=font_style)
                    product_id_button.pack(side="left", pady=5)

                    # 제품 주소 = label
                    label = ctk.CTkLabel(url_frame, text=iherb_url, font=font_style)
                    label.pack(side="left", padx=5, pady=5)

                    return iherb_urls # <--- 제품 주소들이 누적된 배열을 반환합니다!
        # ================================================================================
        '''
        제품 사진을 수집하고, 금지 성분을 확인하는 함수
        '''
        def images_and_ingredients():
            amazon_iherb_value = amazon_iherb_option_var.get()

            log_textbox.delete("0.0", ctk.END)
            answer = "처리 중입니다. 여유롭게 기다려주세요!"
            log_textbox.insert("0.0", answer)

            if amazon_iherb_value == "amazon": # <------------------------------------------------------------------------------------------------------------------------------------------------------

                from selenium.webdriver import ChromeOptions
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service as ChromeService
                from webdriver_manager.chrome import ChromeDriverManager

                options = ChromeOptions()
                options.add_experimental_option("detach", True)
                # options.add_argument("--headless")
                driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

                from selenium.webdriver.common.by import By
                import time

                '''''''''''''''''''''''''''
                아마존 sign-in
                '''''''''''''''''''''''''''
                # 아마존 'sign-in'에 접속합니다.
                driver.get('https://www.amazon.com/-/ko/ap/signin?openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com%2F%3Flanguage%3Dko_KR%26ref_%3Dnav_ya_signin&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=usflex&openid.mode=checkid_setup&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&')

                # # 전체화면으로 실행합니다. 전체화면으로 실행해야 오류가 발생할 확률이 낮아집니다!
                # driver.maximize_window()

                # 아마존 id, password를 입력합니다.
                time.sleep(0.5)
                driver.find_element(By.CSS_SELECTOR, '#ap_email').send_keys(id_entry.get())
                time.sleep(1)
                driver.find_element(By.CSS_SELECTOR, '#continue').click()

                time.sleep(0.5)
                driver.find_element(By.CSS_SELECTOR, '#ap_password').send_keys(password_entry.get())
                time.sleep(1)
                driver.find_element(By.CSS_SELECTOR, '#signInSubmit').click()

                for url in amazon_urls: # <--- 'check_duplicates_and_add_URL()' 함수에서 반환된 배열입니다.
                    '''''''''''''''''''''''''''
                    제품 사진 수집 (반복문)
                    '''''''''''''''''''''''''''
                    # 해당 상품 주소로 이동합니다.
                    time.sleep(1)
                    driver.get(url)

                    # 상품의 이름을 추출합니다.
                    product_name = driver.find_element(By.CSS_SELECTOR, "#productTitle").text

                    # 상품의 'asin_code'를 추출합니다.
                    import re 

                    pattern = r'dp\/([A-Z0-9]{10})\/'
                    asin_code = re.search(pattern, url).group(1)

                    import os

                    # 'amazon' folder 안에 이름이 'asin_code'인 folder를 생성합니다.
                    folder_name = f"{asin_code}"
                    folder_path = os.path.join("amazon", folder_name)
                    os.makedirs(folder_path, exist_ok=True)
                    print(f"{asin_code} # folder created!")

                    # 상품 사진을 click해서 확대창을 띄웁니다.
                    driver.find_element(By.CSS_SELECTOR, '#imgTagWrapperId').click()
                    time.sleep(1)

                    import requests
                    from selenium.common.exceptions import NoSuchElementException
                    from PIL import Image, ImageOps

                    num_a = 0
                    
                    while True:
                        css_selector = f"#ivImage_{num_a} > div" # <--- 해당 제품의 모든 사진(동영상 제외)을 수집합니다.
                        try:
                            # 수집할 사진을 선택합니다.
                            driver.find_element(By.CSS_SELECTOR, css_selector).click()

                            # 선택한 사진의 주소를 추출합니다.
                            time.sleep(2)
                            image_url = driver.find_element(By.CSS_SELECTOR, '#ivLargeImage > img').get_attribute('src')

                            # 사진(원본)을 저장합니다.
                            response = requests.get(image_url)
                            filename = f"image{num_a + 1}.jpg"
                            print(image_url)
                            with open(f"./amazon/{asin_code}/{filename}", 'wb+') as f:
                                f.write(response.content)
                            print(f"{num_a + 1} # save complete!")

                            # 가공할 사진을 불러옵니다.
                            image = Image.open(f"./amazon/{asin_code}/{filename}")

                            '''''''''''''''''''''''''''
                            사진 가공 (네이버)
                            '''''''''''''''''''''''''''
                            if naver_checkbox.get(): # <--- 네이버가 선택되었을 경우입니다!
                                folder_name = "naver(1000, 860)"
                                folder_path = os.path.join(f"amazon/{asin_code}", folder_name)
                                os.makedirs(folder_path, exist_ok=True)

                                # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                new_image = ImageOps.pad(image, (820, 820), color='white')

                                # 흰색 테두리의 두께를 설정합니다. (위, 아래 20)
                                border_thickness = 20

                                # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (860x860)
                                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                # 가공한 사진을 저장합니다.
                                image_with_border.save(f"./amazon/{asin_code}/naver(1000, 860)/{filename}")
                                print(f"{num_a + 1} # retouch(naver) complete!")

                                '''''''''''''''''''''''''''
                                thumbnail 생성 (네이버)
                                '''''''''''''''''''''''''''
                                thumbnail_path_naver = f"./amazon/{asin_code}/naver(1000, 860)/naver_thumbnail.jpg"

                                if os.path.exists(thumbnail_path_naver):
                                    pass # <--- 제품 thumbnail이 이미 존재하는 경우, 구문을 통과합니다.
                                else:
                                    # thumbnail로 만들 사진을 불러옵니다.
                                    image = Image.open(f"./amazon/{asin_code}/naver(1000, 860)/image1.jpg")

                                    # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                    new_image = ImageOps.pad(image, (960, 960), color='white')

                                    # 흰색 테두리의 두께를 설정합니다. (위, 아래 10)
                                    border_thickness = 10

                                    # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (980x980)
                                    image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                    # 색상 테두리의 두께를 설정합니다. (위, 아래 10)
                                    border_thickness = 10

                                    # 테두리 색상을 설정합니다.
                                    border_color = thumbnail_color

                                    # 위에서 설정한 색상 두께를 적용해서 사진 확장을 확장합니다. (1000x1000)
                                    thumbnail_image_for_naver = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                                    # thumbnail을 저장합니다.
                                    thumbnail_image_for_naver.save(f"./amazon/{asin_code}/naver(1000, 860)/naver_thumbnail.jpg")
                                    print("= naver_thumbnail created!")

                            '''''''''''''''''''''''''''
                            사진 가공 (쿠팡)
                            '''''''''''''''''''''''''''
                            if coupang_checkbox.get(): # <--- 쿠팡이 선택되었을 경우입니다!
                                folder_name = "coupang(500, 780)"
                                folder_path = os.path.join(f"amazon/{asin_code}", folder_name)
                                os.makedirs(folder_path, exist_ok=True)

                                # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                new_image = ImageOps.pad(image, (740, 740), color='white')

                                # 흰색 테두리의 두께를 설정합니다. (위, 아래 20)
                                border_thickness = 20

                                # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (780x780)
                                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                # 가공한 사진을 저장합니다.
                                image_with_border.save(f"./amazon/{asin_code}/coupang(500, 780)/{filename}")
                                print(f"{num_a + 1} # retouch(coupang) complete!")

                                '''''''''''''''''''''''''''
                                thumbnail 생성 (쿠팡)
                                '''''''''''''''''''''''''''
                                thumbnail_path_coupang = f"./amazon/{asin_code}/coupang(500, 780)/coupang_thumbnail.jpg"

                                if os.path.exists(thumbnail_path_coupang):
                                    pass
                                else:
                                    # thumbnail로 만들 사진을 불러옵니다.
                                    image = Image.open(f"./amazon/{asin_code}/coupang(500, 780)/image1.jpg")

                                    # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                    new_image = ImageOps.pad(image, (460, 460), color='white')

                                    # 흰색 테두리의 두께를 설정합니다. (위, 아래 10)
                                    border_thickness = 10

                                    # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (480x480)
                                    image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                    # 색상 테두리의 두께를 설정합니다. (위, 아래 10)
                                    border_thickness = 10

                                    # 테두리 색상을 설정합니다.
                                    border_color = thumbnail_color

                                    # 위에서 설정한 색상 두께를 적용해서 사진 확장을 확장합니다. (500x500)
                                    thumbnail_image_for_coupang = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                                    # thumbnail을 저장합니다.
                                    thumbnail_image_for_coupang.save(f"./amazon/{asin_code}/coupang(500, 780)/coupang_thumbnail.jpg")
                                    print("= coupang_thumbnail created!")
                            
                            num_a = num_a + 1

                        except NoSuchElementException:
                            print("<----- no more image ----->")
                            break

                    '''''''''''''''''''''''''''
                    제품 사진 확대창 닫기
                    '''''''''''''''''''''''''''
                    time.sleep(0.5)
                    for num_close in range(0, 10):
                        css_selector = f"#a-popover-{num_close} > div > header > button"
                        try:
                            driver.find_element(By.CSS_SELECTOR, css_selector).click()
                        except NoSuchElementException:
                            pass
                            # print(f"element not found for {css_selector}")
                            
                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "모든 사진이 정상적으로 저장되었습니다.")

                    '''''''''''''''''''''''''''
                    아마존 OCR
                    '''''''''''''''''''''''''''
                    import pytesseract
            
                    warnings = [] # 의심되는 금지 성분이 누적되서 저장될 빈 배열을 선언합니다.

                    num_a = 0

                    while os.path.exists(f"./amazon/{asin_code}/image{num_a + 1}.jpg"):
                
                        ocr_text = pytesseract.image_to_string(Image.open(f"./amazon/{asin_code}/image{num_a + 1}.jpg"), lang='eng')
                        
                        with open('./2022.10.16.txt', 'r') as f:
                            word_list = [line.strip() for line in f.readlines() if line.strip()]
                        
                        # OCR 결과를 검사할 단어 목록(금지 성분 목록표)과 비교하여 겹치는 단어가 있으면 출력합니다.
                        found = False

                        for word in word_list:
                            if word.lower() in ocr_text.lower():
                                ocr_words = ocr_text.split()

                                for ocr_word in ocr_words:
                                    if word.lower() in ocr_word.lower():
                                        ocr_word = ocr_word.replace("(", "").replace(")", "").replace(".", "").replace(",", "")
                                        warning_message = f"[ {ocr_word} ] # {word}"
                                        warning_message = warning_message.replace("?", "").replace("_", "")
                                        
                                        if warning_message not in warnings: # <--- 이전에 추가된 'warning_message'와 중복되지 않는다면,
                                            warnings.append(warning_message) # <--- 배열에 누적해서 저장합니다.

                                        found = True

                                if not found:
                                    pass

                        num_a += 1

                    # 저장된 의심되는 금지 성분 목록(배열)을 출력합니다. 금지 성분이 발견되지 않은 사진은 표시되지 않습니다.
                    # print(f"\n{warnings}\n")

                    # Excel에 저장할 수 있도록 list를 문자열로 변환합니다.
                    warnings_string = "\n".join(warnings)

                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "금지 성분 검사가 완료되었습니다.")

                    '''''''''''''''''''''''''''
                    Excel에 제품 정보 저장
                    '''''''''''''''''''''''''''
                    path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
                    workbook = openpyxl.load_workbook(path)
                    sheet = workbook["amazon"]

                    # 비어있는 행을 식별하고, 삭제합니다.
                    empty_rows = []
                    for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외합니다.
                        if all(cell.value is None for cell in row):
                            empty_rows.append(row)

                    for row in empty_rows:
                        sheet.delete_rows(row[0].row)

                    row_values = [asin_code, product_name, warnings_string, url]

                    print(row_values[0], row_values[1], row_values[2])

                    sheet.append(row_values)
                    workbook.save(path)

                    '''''''''''''''''''''''''''
                    제품 정보 frame 생성
                    '''''''''''''''''''''''''''
                    def display_image(asin_code_for_frame): # <--- 제품 사진을 출력합니다.
                        image_path = f"./amazon/{asin_code_for_frame}/image1.jpg"
                        image = Image.open(image_path)
                        new_image = ImageOps.pad(image, (180, 180), color='white')
                        border_thickness = 10
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                        photo = ImageTk.PhotoImage(image_with_border)
                        product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                        product_image_canvas.image = photo

                    def warnings_string_textbox(asin_code_for_frame): # <--- 의심되는 금지 성분을 출력합니다.
                        product_list = pd.read_excel("./product_list.xlsx", sheet_name="amazon")
                        filtered_row = product_list[product_list.iloc[:, 0] == asin_code_for_frame]
                        value = filtered_row.iloc[0, 2]
                        
                        import numpy as np

                        if not isinstance(value, float) or not np.isnan(value):
                            ingredients_textbox.delete("0.0", ctk.END)
                            ingredients_textbox.insert("0.0", value)
                        else:
                            ingredients_textbox.delete("0.0", ctk.END)
                            ingredients_textbox.insert("0.0", "")

                    def warning_label(asin_code_for_frame): # <--- 금지 성분 존재 여부를 출력합니다.
                        product_list = pd.read_excel("./product_list.xlsx", sheet_name="amazon")
                        filtered_row = product_list[product_list.iloc[:, 0] == asin_code_for_frame]
                        value = filtered_row.iloc[0, 2]

                        import numpy as np

                        if not isinstance(value, float) or not np.isnan(value):
                            warning_value_lable.configure(text="의심되는 성분이 존재합니다!")
                        else:
                            warning_value_lable.configure(text="금지성분이 발견되지 않았습니다.")

                    def asin_code_button_click(asin_code_for_frame): # <--- 'asin_code_button'이 click되었을 때 실행됩니다.
                        print('asin_code_button_click ' + asin_code_for_frame)
                        # 제품 사진
                        display_image(asin_code_for_frame)
                        # 금지 성분
                        warnings_string_textbox(asin_code_for_frame)
                        # 제품 번호
                        code_lable.configure(text=asin_code_for_frame)
                        # 금지 성분 존재 여부
                        warning_label(asin_code_for_frame)

                    def pass_button_click(asin_code_for_frame, frame): # <--- 'pass_button'이 click되었을 때 실행됩니다.
                        print('pass_button_click ' + asin_code_for_frame)

                        frame.configure(fg_color="#217346")

                        log_textbox.delete("0.0", ctk.END)
                        log_textbox.insert("0.0", f"'{asin_code_for_frame}'는 합격입니다.")

                    def fail_button_click(asin_code_for_frame, frame): # <--- 'fail_button'이 click되었을 때 실행됩니다.
                        print('fail_button_click ' + asin_code_for_frame)

                        log_textbox.delete("0.0", ctk.END)
                        log_textbox.insert("0.0", f"'{asin_code_for_frame}'는 불합격입니다!")
                        frame.destroy()

                        path = "./product_list.xlsx"
                        wb = openpyxl.load_workbook(path)
                        sheet = wb["amazon"]

                        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                            if row[0].value == asin_code_for_frame:
                                sheet.delete_rows(row[0].row)
                                break

                        wb.save(path)

                        # 사진이 들어있는 folder를 삭제합니다.
                        import shutil

                        folder_path = f"./amazon/{asin_code_for_frame}"
                        if os.path.exists(folder_path):
                            shutil.rmtree(folder_path)

                    # 제품 정보 = frame
                    product_frame = ctk.CTkFrame(scrollable_frame)
                    product_frame.pack(fill="x", pady=(5,0))

                    # 제품 번호(asin_code) = button
                    asin_code_button = ctk.CTkButton(product_frame, text=f"{asin_code}", width=50, font=font_style, 
                        command=lambda asin_code_for_frame=asin_code: 
                        asin_code_button_click(asin_code_for_frame))
                    asin_code_button.pack(side="left", padx=(5,0), pady=5)

                    # 합격 = button
                    pass_button = ctk.CTkButton(product_frame, text="pass", width=50, fg_color="#217346", hover_color="#005000", font=font_style, 
                        command=lambda asin_code_for_frame=asin_code, 
                        frame=product_frame: 
                        pass_button_click(asin_code_for_frame, frame))
                    pass_button.pack(side="left", padx=5, pady=5)

                    # 불합격 = button
                    fail_button = ctk.CTkButton(product_frame, text="fail", width=50, fg_color="#CC3D3D", hover_color="#960707", font=font_style,
                        command=lambda asin_code_for_frame=asin_code, 
                        frame=product_frame: 
                        fail_button_click(asin_code_for_frame, frame))
                    fail_button.pack(side="left", pady=5)

                    # 제품명 = label
                    label = ctk.CTkLabel(product_frame, text=product_name, font=font_style)
                    label.pack(side="left", padx=5, pady=5, anchor="center")

                # web browser를 종료합니다.
                driver.quit()

                log_textbox.delete("0.0", ctk.END)
                log_textbox.insert("0.0", "모든 기능이 정상적으로 실행 완료되었습니다!")

            else : # <------------------------------------------------------------------------------------------------------------------------------------------------------
                import os
                import requests
                from bs4 import BeautifulSoup

                for url in iherb_urls: # <--- 'check_duplicates_and_add_URL()' 함수에서 반환된 배열입니다.
                    '''''''''''''''''''''''''''
                    제품 사진 수집 (반복문)
                    '''''''''''''''''''''''''''
                    # 상품의 'product_id'를 추출합니다.
                    import re
            
                    product_id = url.rsplit('/', 1)[-1].replace("?rec=home", "")
                    # product_id = re.search(r"/(\d+)$", url).group(1)

                    # 'iherb' folder 안에 이름이 'product_id'인 folder를 생성합니다.
                    folder_name = f"{product_id}"
                    folder_path = os.path.join("iherb", folder_name)
                    os.makedirs(folder_path, exist_ok=True)
                    print(f"{product_id} # folder created!")

                    headers = { # <----- 사용자마다 다르게 설정해야 합니다!
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
                    }

                    response = requests.get(url, headers=headers)
                    soup = BeautifulSoup(response.content, "html.parser")

                    img_tags = soup.find_all("img", {"data-large-img": True})

                    from PIL import Image, ImageOps

                    for num_i, img_tag in enumerate(img_tags):
                        img_url = img_tag["data-large-img"]

                        # 사진의 이름을 설정합니다.
                        filename = f"image{num_i + 1}.jpg"

                        # 제품 사진을 받아오기 위한 요청을 보냅니다.
                        img_response = requests.get(img_url, headers=headers)

                        # 사진(원본)을 저장합니다.
                        with open(f"./iherb/{product_id}/{filename}", 'wb+') as f:
                                f.write(img_response.content)
                        print(f"{num_i + 1} # save complete!")

                        # 가공할 사진을 불러옵니다.
                        image = Image.open(f"./iherb/{product_id}/{filename}")

                        '''''''''''''''''''''''''''
                        사진 가공 (네이버)
                        '''''''''''''''''''''''''''
                        if naver_checkbox.get():
                            folder_name = "naver(1000, 860)"
                            folder_path = os.path.join(f"iherb/{product_id}", folder_name)
                            os.makedirs(folder_path, exist_ok=True)

                            # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                            new_image = ImageOps.pad(image, (820, 820), color='white')

                            # 흰색 테두리의 두께를 설정합니다. (위, 아래 20)
                            border_thickness = 20

                            # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (860x860)
                            image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')           

                            # 가공한 사진을 저장합니다.
                            image_with_border.save(f"./iherb/{product_id}/naver(1000, 860)/{filename}")
                            print(f"{num_i + 1} # retouch(naver) complete!")

                            '''''''''''''''''''''''''''
                            thumbnail 생성 (네이버)
                            '''''''''''''''''''''''''''
                            thumbnail_path_naver = f"./iherb/{product_id}/naver(1000, 860)/naver_thumbnail.jpg"

                            if os.path.exists(thumbnail_path_naver):
                                pass
                            else:
                                # thumbnail로 만들 사진을 불러옵니다.
                                image = Image.open(f"./iherb/{product_id}/naver(1000, 860)/image1.jpg")

                                # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                new_image = ImageOps.pad(image, (960, 960), color='white')

                                # 흰색 테두리의 두께를 설정합니다. (위, 아래 10)
                                border_thickness = 10

                                # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (980x980)
                                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                # 색상 테두리의 두께를 설정합니다. (위, 아래 10)
                                border_thickness = 10

                                # 테두리 색상을 설정합니다.
                                border_color = thumbnail_color

                                # 위에서 설정한 색상 두께를 적용해서 사진 확장을 확장합니다. (1000x1000)
                                thumbnail_image_for_naver = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                                # thumbnail을 저장합니다.
                                thumbnail_image_for_naver.save(f"./iherb/{product_id}/naver(1000, 860)/naver_thumbnail.jpg")
                                print("= naver_thumbnail created!")

                        '''''''''''''''''''''''''''
                        사진 가공 (쿠팡)
                        '''''''''''''''''''''''''''
                        if coupang_checkbox.get():
                            folder_name = "coupang(500, 780)"
                            folder_path = os.path.join(f"iherb/{product_id}", folder_name)
                            os.makedirs(folder_path, exist_ok=True)

                            # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                            new_image = ImageOps.pad(image, (740, 740), color='white')

                            # 흰색 테두리의 두께를 설정합니다. (위, 아래 20)
                            border_thickness = 20

                            # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (780x780)
                            image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                            # 가공한 사진을 저장합니다.
                            image_with_border.save(f"./iherb/{product_id}/coupang(500, 780)/{filename}")
                            print(f"{num_i + 1} # retouch(coupang) complete!")

                            '''''''''''''''''''''''''''
                            thumbnail 생성 (쿠팡)
                            '''''''''''''''''''''''''''
                            thumbnail_path_coupang = f"./iherb/{product_id}/coupang(500, 780)/coupang_thumbnail.jpg"

                            if os.path.exists(thumbnail_path_coupang):
                                pass
                            else:
                                # thumbnail로 만들 사진을 불러옵니다.
                                image = Image.open(f"./iherb/{product_id}/coupang(500, 780)/image1.jpg")

                                # 사진 크기를 조정하고, 부족한 부분을 흰색으로 채웁니다.
                                new_image = ImageOps.pad(image, (460, 460), color='white')

                                # 흰색 테두리 두께 설정
                                border_thickness = 10

                                # 위에서 설정한 두께를 적용해서 사진 확장을 확장합니다. (480x480)
                                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                                # 색상 테두리의 두께를 설정합니다. (위, 아래 10)
                                border_thickness = 10

                                # 테두리 색상을 설정합니다.
                                border_color = thumbnail_color

                                # 위에서 설정한 색상 두께를 적용해서 사진 확장을 확장합니다. (500x500)
                                thumbnail_image_for_coupang = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                                # thumbnail을 저장합니다.
                                thumbnail_image_for_coupang.save(f"./iherb/{product_id}/coupang(500, 780)/coupang_thumbnail.jpg")
                                print("= coupang_thumbnail created!")
                    
                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "모든 사진이 정상적으로 저장되었습니다.")

                    # 상품 이름 추출
                    product_name = soup.find('div', class_='product-summary-title').get_text(strip=True)
                    print(product_name)

                    '''''''''''''''''''''''''''
                    아이허브 OCR
                    '''''''''''''''''''''''''''
                    warnings = [] # 의심되는 금지 성분이 누적되서 저장될 빈 배열을 선언합니다.

                    # 영양 성분 정보
                    container = soup.find('div', {'class': 'supplement-facts-container'})
                    supplement_facts_text = container.get_text(strip=True)
                    supplement_facts_text = supplement_facts_text.replace('영양 성분 정보', '영양 성분 정보\n')

                    '''

                    # 상품 사용법
                    container = soup.find('div', {'class': 'prodOverviewDetail'})
                    suggested_use_text = container.get_text(strip=True)
                    print("----- 상품 사용법 -----")
                    print(suggested_use_text)

                    '''

                    # 포함된 다른 성분들
                    container = soup.find('div', {'class': 'prodOverviewIngred'})
                    other_ingredients_text = container.get_text(strip=True)
                    other_ingredients_text = other_ingredients_text.replace('주요 성분', '주요 성분\n')
                    other_ingredients_text = other_ingredients_text.replace('기타 성분', '\n기타 성분\n')

                    if '무함유.' in other_ingredients_text:
                        other_ingredients_text = '무함유.'
                    elif '이 제품은' in other_ingredients_text:
                        other_ingredients_text = other_ingredients_text.split('이 제품은')[0]
                    elif '이 제품에는' in other_ingredients_text:
                        other_ingredients_text = other_ingredients_text.split('이 제품에는')[0]
                    else:
                        pass
                        
                    from googletrans import Translator
                    translator = Translator()
                    import time
                    time.sleep(0.5)

                    # 'supplement_facts_text'에서 '%하루 영양소 기준치' 이후의 문자열을 추출합니다.
                    match = re.search(r'%하루 영양소 기준치(.*?)$', supplement_facts_text, re.DOTALL)
                    if match:
                        extracted_text = match.group(1).strip()
                    supplement_facts_text = extracted_text

                    # 번역할 문자열입니다.
                    text_to_translate = supplement_facts_text + " " + other_ingredients_text

                    # 문자열을 영어로 번역합니다.
                    translated_text = translator.translate(text_to_translate, dest='en').text
                    translated_text = translated_text.replace("*", "").replace("The standard value per day is not set.", ", ")
                    translated_text = translated_text.replace("Nothing.", "")

                    # 번역문을 출력합니다.
                    # print(f"\n{translated_text}\n")

                    # 검사할 단어 목록을 메모장에서 불러옵니다.
                    with open('./2022.10.16.txt', 'r') as f:
                        word_list = [line.strip() for line in f.readlines() if line.strip()]

                    # 영어 번역문을 검사할 단어 목록(금지 성분 목록표)과 비교하여 겹치는 단어가 있으면 출력합니다.
                    found = False

                    for word in word_list:
                        if word.lower() in translated_text.lower():
                            ocr_words = translated_text.split()

                            for ocr_word in ocr_words:
                                if word.lower() in ocr_word.lower():
                                    ocr_word = ocr_word.replace("(", "").replace(")", "").replace(".", "").replace(",", "")
                                    warning_message = f"[ {ocr_word} ] # {word}"
                                    warning_message = warning_message.replace("?", "").replace("_", "")

                                    if warning_message not in warnings: # <--- 이전에 추가된 'warning_message'와 중복되지 않는다면,
                                        warnings.append(warning_message) # <--- 배열에 누적해서 저장합니다.

                                    found = True

                            if not found:
                                pass

                    # 저장된 의심되는 금지 성분 목록(배열)을 출력합니다. 금지 성분이 발견되지 않은 사진은 표시되지 않습니다.
                    # print(f"\n{warnings}\n")

                    # Excel에 저장할 수 있도록 list를 문자열로 변환합니다.
                    warnings_string = "\n".join(warnings)

                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "금지 성분 검사가 완료되었습니다.")

                    '''''''''''''''''''''''''''
                    Excel에 제품 정보 저장
                    '''''''''''''''''''''''''''
                    path = "./product_list.xlsx"
                    workbook = openpyxl.load_workbook(path)
                    sheet = workbook["iherb"]

                    # 비어있는 행을 식별하고, 삭제합니다.
                    empty_rows = []
                    for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외합니다.
                        if all(cell.value is None for cell in row):
                            empty_rows.append(row)

                    for row in empty_rows:
                        sheet.delete_rows(row[0].row)

                    row_values = [product_id, product_name, warnings_string, url]

                    print(row_values[0], row_values[1], row_values[2])

                    sheet.append(row_values)
                    workbook.save(path)

                    '''''''''''''''''''''''''''
                    제품 정보 frame 생성
                    '''''''''''''''''''''''''''
                    def iherb_display_image(product_id_for_frame): # <--- 제품 사진을 출력합니다.
                        image_path = f"./iherb/{product_id_for_frame}/image1.jpg"
                        image = Image.open(image_path)
                        new_image = ImageOps.pad(image, (180, 180), color='white')
                        border_thickness = 10
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                        photo = ImageTk.PhotoImage(image_with_border)
                        product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                        product_image_canvas.image = photo

                    def iherb_warnings_string_textbox(product_id_for_frame): # <--- 의심되는 금지 성분을 출력합니다.
                        iherb_product_list = pd.read_excel("./product_list.xlsx", sheet_name="iherb")
                        iherb_filtered_row = iherb_product_list[iherb_product_list.iloc[:, 0].astype(str) == str(product_id_for_frame)]
                        iherb_value = iherb_filtered_row.iloc[0, 2]
                        
                        import numpy as np

                        if not isinstance(iherb_value, float) or not np.isnan(iherb_value):
                            ingredients_textbox.delete("0.0", ctk.END)
                            ingredients_textbox.insert("0.0", iherb_value)
                        else:
                            ingredients_textbox.delete("0.0", ctk.END)
                            ingredients_textbox.insert("0.0", "")

                    def iherb_warning_label(product_id_for_frame): # <--- 금지 성분 존재 여부를 출력합니다.
                        iherb_product_list = pd.read_excel("./product_list.xlsx", sheet_name="iherb")
                        iherb_filtered_row = iherb_product_list[iherb_product_list.iloc[:, 0].astype(str) == str(product_id_for_frame)]
                        iherb_value = iherb_filtered_row.iloc[0, 2]

                        import numpy as np

                        if not isinstance(iherb_value, float) or not np.isnan(iherb_value):
                            warning_value_lable.configure(text="의심되는 성분이 존재합니다!")
                        else:
                            warning_value_lable.configure(text="금지성분이 발견되지 않았습니다.")

                    def product_id_button_click(product_id_for_frame): # <--- 'product_id_button'이 click되었을 때 실행됩니다.
                        print('product_id_button_click ' + product_id_for_frame)
                        # 제품 사진
                        iherb_display_image(product_id_for_frame)
                        # 금지 성분
                        iherb_warnings_string_textbox(product_id_for_frame)
                        # 제품 번호
                        code_lable.configure(text=product_id_for_frame)
                        # 금지 성분 존재 여부
                        iherb_warning_label(product_id_for_frame)

                    def iherb_pass_button_click(product_id_for_frame, frame): # <--- 'pass_button'이 click되었을 때 실행됩니다.
                        print('iherb_pass_button_click ' + product_id_for_frame)

                        frame.configure(fg_color="#217346")

                        log_textbox.delete("0.0", ctk.END)
                        log_textbox.insert("0.0", f"'{product_id_for_frame}'는 합격입니다.")

                    def iherb_fail_button_click(product_id_for_frame, frame): # <--- 'fail_button'이 click되었을 때 실행됩니다.
                        print('iherb_fail_button_click ' + product_id_for_frame)

                        log_textbox.delete("0.0", ctk.END)
                        log_textbox.insert("0.0", f"'{product_id_for_frame}'는 불합격입니다!")
                        frame.destroy()

                        path = "./product_list.xlsx"
                        wb = openpyxl.load_workbook(path)
                        sheet = wb["iherb"]

                        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                            if row[0].value == product_id_for_frame:
                                sheet.delete_rows(row[0].row)
                                break

                        wb.save(path)

                        # 사진이 들어있는 folder를 삭제합니다.
                        import shutil

                        folder_path = f"./iherb/{product_id_for_frame}"
                        if os.path.exists(folder_path):
                            shutil.rmtree(folder_path)

                    # 제품 정보 = frame
                    product_frame = ctk.CTkFrame(scrollable_frame)
                    product_frame.pack(fill="x", pady=(5,0))

                    # 제품 번호(product_id) = button
                    product_id_button = ctk.CTkButton(product_frame, text=f"{product_id}", width=50, font=font_style, 
                        command=lambda product_id_for_frame=product_id:
                        product_id_button_click(product_id_for_frame))
                    product_id_button.pack(side="left", padx=(5,0), pady=5)

                    # 합격 = button
                    pass_button = ctk.CTkButton(product_frame, text="pass", width=50, fg_color="#217346", hover_color="#005000", font=font_style, 
                        command=lambda product_id_for_frame=product_id, 
                        frame=product_frame: 
                        iherb_pass_button_click(product_id_for_frame, frame))
                    pass_button.pack(side="left", padx=5, pady=5)

                    # 불합격 = button
                    fail_button = ctk.CTkButton(product_frame, text="fail", width=50, fg_color="#CC3D3D", hover_color="#960707", font=font_style, 
                        command=lambda product_id_for_frame=product_id, 
                        frame=product_frame: 
                        iherb_fail_button_click(product_id_for_frame, frame))
                    fail_button.pack(side="left", pady=5)

                    # 제품명 = label
                    label = ctk.CTkLabel(product_frame, text=product_name, font=font_style)
                    label.pack(side="left", padx=5, pady=5, anchor="center")

                log_textbox.delete("0.0", ctk.END)
                log_textbox.insert("0.0", "모든 기능이 정상적으로 실행 완료되었습니다!")
        # ================================================================================
        '''
        excel 자료(product_list - amazon, iherb)의 비어있는 행들을 삭제하고 최종적으로 저장하는 함수
        '''
        def save_to_a_database():
            amazon_iherb_value = amazon_iherb_option_var.get()

            if amazon_iherb_value == "amazon":
                path = "./product_list.xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook["amazon"]

                # 비어있는 행을 식별하고, 삭제합니다.
                empty_rows = []
                for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외합니다.
                    if all(cell.value is None for cell in row):
                        empty_rows.append(row)

                for row in empty_rows:
                    sheet.delete_rows(row[0].row)

                workbook.save(path)

            else:
                path = "./product_list.xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook["iherb"]

                # 비어있는 행을 식별하고, 삭제합니다.
                empty_rows = []
                for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외합니다.
                    if all(cell.value is None for cell in row):
                        empty_rows.append(row)

                for row in empty_rows:
                    sheet.delete_rows(row[0].row)

                workbook.save(path)

            root.destroy()
        # ================================================================================
        '''
        제품 번호를 입력했을 때, 제품 정보와 제품 주소(URL)를 출력하는 함수 
        '''
        def search():
            amazon_iherb_value = amazon_iherb_option_var.get()

            if amazon_iherb_value == "amazon": 
                search_asin_code = code_entry.get()
                search_amazon_product_list = pd.read_excel("./product_list.xlsx", sheet_name="amazon")
                filtered_row = search_amazon_product_list[search_amazon_product_list.iloc[:, 0] == search_asin_code]

                if filtered_row.empty or filtered_row is False:
                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "[경고] 해당 제품은 존재하지 않습니다!")

                else:
                    search_amazon_value = filtered_row.iloc[0, 3]

                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "제품 검색 완료.")
                    
                    url_textbox.delete("0.0", ctk.END)
                    url_textbox.insert("0.0", search_amazon_value)
                
                ''''''

                from PIL import Image, ImageOps

                img_path=f"./amazon/{search_asin_code}/image1.jpg"
                image = Image.open(img_path)
                new_image = ImageOps.pad(image, (180, 180), color='white')
                border_thickness = 10
                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                photo = ImageTk.PhotoImage(image_with_border)
                product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                product_image_canvas.image = photo

                ''''''

                ingredients_filtered_row = search_amazon_product_list[search_amazon_product_list.iloc[:, 0] == search_asin_code]
                value = ingredients_filtered_row.iloc[0, 2]
                        
                import numpy as np

                if not isinstance(value, float) or not np.isnan(value):
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", value)
                else:
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", "")

                ''''''

                warning_filtered_row = search_amazon_product_list[search_amazon_product_list.iloc[:, 0] == search_asin_code]
                warning_value = warning_filtered_row.iloc[0, 2]

                import numpy as np

                if not isinstance(value, float) or not np.isnan(warning_value):
                    warning_value_lable.configure(text="의심되는 성분이 존재합니다!")
                else:
                    warning_value_lable.configure(text="금지성분이 발견되지 않았습니다.")

                ''''''
                code_lable.configure(text=search_asin_code)

            else:
                search_product_id = code_entry.get()
                search_iherb_product_list = pd.read_excel("./product_list.xlsx", sheet_name="iherb")
                filtered_row = search_iherb_product_list[search_iherb_product_list.iloc[:, 0].astype(str) == str(search_product_id)]
                
                if filtered_row.empty or filtered_row is False:
                    log_textbox.delete("0.0", ctk.END)
                    log_textbox.insert("0.0", "[경고] 해당 제품은 존재하지 않습니다!")

                else:
                    search_iherb_value = filtered_row.iloc[0, 3]

                    url_textbox.delete("0.0", ctk.END)
                    url_textbox.insert("0.0", search_iherb_value)

                ''''''

                from PIL import Image, ImageOps

                img_path=f"./iherb/{search_product_id}/image1.jpg"
                image = Image.open(img_path)
                new_image = ImageOps.pad(image, (180, 180), color='white')
                border_thickness = 10
                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                photo = ImageTk.PhotoImage(image_with_border)
                product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                product_image_canvas.image = photo

                ''''''

                ingredients_filtered_row = search_iherb_product_list[search_iherb_product_list.iloc[:, 0].astype(str) == str(search_product_id)]
                value = ingredients_filtered_row.iloc[0, 2]
                        
                import numpy as np

                if not isinstance(value, float) or not np.isnan(value):
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", value)
                else:
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", "")

                ''''''

                warning_filtered_row = search_iherb_product_list[search_iherb_product_list.iloc[:, 0].astype(str) == str(search_product_id)]
                warning_value = warning_filtered_row.iloc[0, 2]

                import numpy as np

                if not isinstance(value, float) or not np.isnan(warning_value):
                    warning_value_lable.configure(text="의심되는 성분이 존재합니다!")
                else:
                    warning_value_lable.configure(text="금지성분이 발견되지 않았습니다.")

                ''''''
                code_lable.configure(text=search_product_id)
        # ================================================================================
        '''
        제품 주소를 'ctrl + c' 해주는 함수 
        '''
        def copy_URL():
            import pyperclip

            url_text = url_textbox.get("0.0", ctk.END)
            pyperclip.copy(url_text)

            log_textbox.delete("0.0", ctk.END)
            log_textbox.insert("0.0", "주소 복사 완료.")
        # ================================================================================
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        # ui
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        # ================================================================================
        root = ctk.CTk()
        root.title("KME Scraper")

        root.protocol("WM_DELETE_WINDOW", open_toplevel) # 창 종료를 눌렀을 때, 'open_toplevel' 함수를 실행합니다.
        # ================================================================================

        ctk.set_appearance_mode("dark") # system, dark, light
        ctk.set_default_color_theme("blue") # blue(standard), green, dark-blue

        font_style = ctk.CTkFont("돋움", size=13) # 'font=font_style'을 적용하면 글씨체가 '돋움'으로 변합니다.

        # ================================================================================
        frame_01 = ctk.CTkFrame(root)
        frame_01.grid(row=0, column=0, padx=(20,10), pady=(20,10), sticky="news")
        # ================================================================================

        option_frame = ctk.CTkFrame(frame_01)
        option_frame.pack(side="left", fill="both", expand=True, padx=(20,10), pady=20)

        # 아마존 / 아이허브 = radiobutton
        amazon_iherb_option_var = ctk.StringVar(value="amazon") # <--- 분기점의 시초가 되는 중요한 변수입니다!

        amazon_radiobutton = ctk.CTkRadioButton(option_frame, text=" 아마존", fg_color="#EDD200", 
                        variable=amazon_iherb_option_var, value="amazon", font=font_style)
        amazon_radiobutton.pack(side="top", fill="both", expand=True, padx=20, pady=(20,5))
        iherb_radiobutton = ctk.CTkRadioButton(option_frame, text=" 아이허브", fg_color="#22741C", 
                        variable=amazon_iherb_option_var, value="iherb", font=font_style)
        iherb_radiobutton.pack(side="bottom", fill="both", expand=True, padx=20, pady=(5,20))

        ''''''

        entry_frame = ctk.CTkFrame(frame_01)
        entry_frame.pack(fill="both", side="right", expand=True, padx=(10,20), pady=20)

        # Mobile phone number or email = entry
        id_entry = ctk.CTkEntry(entry_frame, placeholder_text=" Mobile phone number or email", font=font_style) # <--- 입력창
        id_entry.pack(fill="x", expand=True, side="top", padx=20, pady=(20,5))

        # Password = entry
        password_entry = ctk.CTkEntry(entry_frame, placeholder_text=" Password", font=font_style) # <--- 입력창
        password_entry.pack(fill="x", expand=True, side="bottom", padx=20, pady=(5,20))

        # ================================================================================
        frame_02 = ctk.CTkFrame(root)
        frame_02.grid(row=1, column=0, padx=(20,10), pady=10, sticky="news")
        # ================================================================================

        image_frame = ctk.CTkFrame(frame_02)
        image_frame.pack(side="top", fill="both", expand=True, padx=20, pady=(20,10))

        # 네이버 = checkbox
        naver_checkbox = ctk.CTkCheckBox(image_frame, text=" 네이버 (1000 , 860)", font=font_style)
        naver_checkbox.pack(expand=True, side="left", padx=(50,10), pady=10)
        # 쿠팡 = checkbox
        coupang_checkbox = ctk.CTkCheckBox(image_frame, text=" 쿠팡 (500 , 780)", font=font_style)
        coupang_checkbox.pack(expand=True, side="left", padx=(10,50), pady=10)

        ''''''

        border_frame = ctk.CTkFrame(frame_02)
        border_frame.pack(side="top", fill="both", expand=True, padx=20, pady=(0,20))

        # 색상 결정 = segemented button
        color_segemented_button = ctk.CTkSegmentedButton(border_frame, font=font_style, 
                        values=[" White ", " Red ", " Orange ", " Yellow ", " Green ", " Blue ", " Purple "], 
                        command=color_segmented_button_callback)
        color_segemented_button.set(" White ")
        color_segemented_button.pack(expand=True, padx=10, pady=10)

        # ================================================================================
        frame_03 = ctk.CTkFrame(root)
        frame_03.grid(row=0, rowspan=3, column=1, padx=(10,20), pady=(20,10), sticky="news")
        # ================================================================================
        # 제품 주소 = entry
        url_entry = ctk.CTkEntry(frame_03, placeholder_text=" 제품 주소", font=font_style) # <--- 입력창
        url_entry.pack(fill="x", padx=20, pady=(20,5))

        url_scrollable_frame = ctk.CTkScrollableFrame(frame_03, width=500)
        url_scrollable_frame.pack(fill="y", expand=True, padx=20, pady=(5,10))

        # 제품 주소(URL) 중복 검사 => 목록 추가 = button
        add_button = ctk.CTkButton(frame_03, width=200, text="제품 주소 중복 검사 => 목록 추가", font=font_style,  
                        command=check_duplicates_and_add_URL)
        add_button.pack(fill="x", padx=20, pady=(0,5))

        # 제품 사진 수집 => 금지 성분 조사 = button
        add_button = ctk.CTkButton(frame_03, width=200, text="제품 사진 수집 => 금지 성분 조사", font=font_style,  
                        command=images_and_ingredients)
        add_button.pack(fill="x", padx=20, pady=(5,20))

        # ================================================================================
        textbox_frame = ctk.CTkFrame(root)
        textbox_frame.grid(row=2, column=0, padx=(20,10), pady=10, sticky="ews")
        # ================================================================================

        # 설명 출력창 = textbox
        log_textbox = ctk.CTkTextbox(textbox_frame, height=30, font=font_style)
        log_textbox.pack(fill="both", expand=True, padx=20, pady=10)

        # ================================================================================
        frame_04 = ctk.CTkFrame(root)
        frame_04.grid(row=3, column=0, padx=(20,10), pady=(10,10), sticky="news")
        # ================================================================================

        # 제품 사진 출력창 = canvas
        product_image_canvas = tk.Canvas(frame_04, width=200, height=200)
        product_image_canvas.grid(row=0, column=0, padx=(20, 10), pady=(20,10))

        ''''''

        ingredients_textbox_frame = ctk.CTkFrame(frame_04)
        ingredients_textbox_frame.grid(row=0, column=1, padx=(10,20), pady=(20,10), sticky="news")

        # 금지 성분 출력창 = textbox
        ingredients_textbox = ctk.CTkTextbox(ingredients_textbox_frame, activate_scrollbars=False, font=font_style)
        ingredients_textbox.pack(side="left", expand=True, fill="both", padx=(5,0), pady=5)

        ctk_textbox_scrollbar = ctk.CTkScrollbar(ingredients_textbox_frame, command=ingredients_textbox.yview)
        ctk_textbox_scrollbar.pack(side="right", fill="y")

        ingredients_textbox.configure(yscrollcommand=ctk_textbox_scrollbar.set)

        ''''''

        product_code = ctk.CTkFrame(frame_04, height=50)
        product_code.grid(row=1, column=0, padx=(20,10), pady=(0,20), sticky="ew")

        # 제품 식별자 출력창 = label
        code_lable = ctk.CTkLabel(product_code, text="", font=font_style)
        code_lable.pack(padx=10, pady=10)

        ''''''

        pass_fail_frame = ctk.CTkFrame(frame_04, height=50)
        pass_fail_frame.grid(row=1, column=1, padx=(10,20), pady=(0,20), sticky="ew")

        # 금지 성분 존재 여부 출력창 = label
        warning_value_lable = ctk.CTkLabel(pass_fail_frame, text="", width=250, font=font_style)
        warning_value_lable.pack(padx=10, pady=10)

        # ================================================================================
        frame_05 = ctk.CTkFrame(root)
        frame_05.grid(row=4, column=0, padx=(20,10), pady=10, sticky="news")
        # ================================================================================

        license_frame = ctk.CTkFrame(frame_05)
        license_frame.pack(side="left", fill="x", expand=True, padx=(20,5), pady=10)

        # 사용자 정보 = label
        license_lable = ctk.CTkLabel(license_frame, text="DESKTOP-LNMR5Q3", font=font_style)
        license_lable.pack(padx=5)

        ''''''

        list_date_frame = ctk.CTkFrame(frame_05)
        list_date_frame.pack(side="left", fill="x", expand=True, padx=5, pady=10)

        # 금지 성분 최신화 = label
        list_date_lable = ctk.CTkLabel(list_date_frame, text="금지 성분 최신화. 2022-10-16", font=font_style)
        list_date_lable.pack(padx=5)

        ''''''

        version_frame = ctk.CTkFrame(frame_05)
        version_frame.pack(side="left", fill="x", expand=True, padx=(5,20), pady=10)

        # KME. 23.1.0 = label
        version_lable = ctk.CTkLabel(version_frame, text="KME 23.1.0", font=font_style)
        version_lable.pack(padx=5)

        # ================================================================================
        frame_06 = ctk.CTkFrame(root)
        frame_06.grid(row=3, rowspan=2, column=1, padx=(10,20), pady=10, sticky="news")
        # ================================================================================

        scrollable_frame = ctk.CTkScrollableFrame(frame_06, width=500)
        scrollable_frame.pack(fill="y", expand=True, padx=20, pady=(20,10))

        # 모든 과정을 완료하셨다면, 눌러주세요! [ 저장 => 종료 ] = button
        save_button = ctk.CTkButton(frame_06, width=200, text="모든 과정을 완료하셨다면 눌러주세요! [ 저장 => 종료 ]", font=font_style, 
                        command=save_to_a_database)
        save_button.pack(fill="x", padx=20, pady=(0,20))

        # ================================================================================
        frame_07 = ctk.CTkFrame(root)
        frame_07.grid(row=5, column=0, columnspan=2, padx=20, pady=(10,20), sticky="news")
        # ================================================================================

        # 제품 번호 = entry
        code_entry = ctk.CTkEntry(frame_07, placeholder_text=" 제품 번호", font=font_style) # <--- 입력창
        code_entry.pack(fill="x", expand=True, side="left", padx=(20,5), pady=10)

        # 검색 = button
        search_button = ctk.CTkButton(frame_07, text="검색", font=font_style, 
                        command=search)
        search_button.pack(fill="x", expand=True, side="left", padx=5, pady=10)

        # 제품 주소(URL) 출력창 = textbox
        url_textbox = ctk.CTkTextbox(frame_07, width=600, height=30, font=font_style)
        url_textbox.pack(fill="x", expand=True, side="left", padx=5, pady=10)

        # 제품 주소 복사 = button
        copy_url_button = ctk.CTkButton(frame_07, text="제품 주소 복사", font=font_style, 
                        command=copy_URL)
        copy_url_button.pack(fill="x", expand=True, side="left", padx=(5,20), pady=10)

        ''''''

        root.mainloop()
    else:
        print('사용 기간이 만료되었습니다!')
else:
    print('')
    print('인증된 사용자가 아닙니다!')
