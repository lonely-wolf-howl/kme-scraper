import customtkinter as ctk
import openpyxl
import tkinter as tk
from tkinter import ttk
import re
from PIL import Image # to create an image object
from PIL import ImageTk # to use an image in the tkinter canvas
import sys
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd

''''''

import subprocess
import pkg_resources

installed_packages = [package.project_name for package in pkg_resources.working_set]
if 'googletrans' in installed_packages:
    package_name = 'googletrans'
    current_version = next((package.version for package in pkg_resources.working_set if package.project_name == package_name), None)
    if current_version and current_version != '4.0.0-rc1':
        subprocess.check_call(['pip', 'install', '--upgrade', 'googletrans==4.0.0-rc1'], 
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
# function
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
amazon_urls = []
iherb_urls = []

thumbnail_color = 'white'

url_frame_list = []
# ================================================================================
def load_data():
    path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
    workbook = pd.read_excel(path, sheet_name='amazon')

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)

    pd.set_option('display.width', None)

    print(workbook)
# ================================================================================
def color_segmented_button_callback(value):
    global thumbnail_color
    thumbnail_color = value.strip()
# ================================================================================
def check_duplicates_and_add_URL():
    amazon_iherb_value = amazon_iherb_option_var.get()

    if amazon_iherb_value == "amazon":
        amazon_url = url_entry.get().strip()
        wb = openpyxl.load_workbook("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx")
        sheet = wb['amazon']
        product_list_amazon_urls = []
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            product_list_amazon_urls.append(row[0].value)

        log_textbox.delete("0.0", ctk.END)
        if amazon_url in product_list_amazon_urls:
            answer = "Warning! duplicate URL found."
            log_textbox.insert("0.0", answer)
        else:
            answer = "Success."
            log_textbox.insert("0.0", answer)

            amazon_urls.append(amazon_url)
            print(amazon_urls)

            # url frame
            url_frame = ctk.CTkFrame(url_scrollable_frame)
            url_frame.pack(fill="x", pady=(5,0))

            # delete button
            delete_button = ctk.CTkButton(url_frame, text="delete", width=50, fg_color="#CC3D3D", hover_color="#960707", 
                                command=lambda frame=url_frame, url=amazon_url: 
                                (amazon_urls.remove(url), log_textbox.delete("0.0", ctk.END), log_textbox.insert("0.0", "Delete complete."), frame.destroy()))
            delete_button.pack(side="left", padx=5, pady=5)

            # asin code
            asin_code = re.search(r"dp\/([A-Z0-9]{10})\/", amazon_url).group(1)
            asin_code_button = ctk.CTkButton(url_frame, text=f"{asin_code}", width=50)
            asin_code_button.pack(side="left", pady=5)

            # url label
            label = ctk.CTkLabel(url_frame, text=amazon_url)
            label.pack(side="left", padx=5, pady=5, anchor="center")

            return amazon_urls

    else:
        iherb_url = url_entry.get().strip()
        wb = openpyxl.load_workbook("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx")
        sheet = wb['iherb']
        product_list_iherb_urls = []
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            product_list_iherb_urls.append(row[0].value)

        log_textbox.delete("0.0", ctk.END)
        if iherb_url in product_list_iherb_urls:
            answer = "Warning! duplicate URL found."
            log_textbox.insert("0.0", answer)
        else:
            answer = "Success!"
            log_textbox.insert("0.0", answer)

            iherb_urls.append(iherb_url)
            # print(iherb_urls)

            url_frame = ctk.CTkFrame(url_scrollable_frame)
            url_frame.pack(fill="x", pady=(5,0))

            delete_button = ctk.CTkButton(url_frame, text="delete", width=50, fg_color="#CC3D3D", hover_color="#960707", 
                                command=lambda frame=url_frame, url=iherb_url: 
                                (iherb_urls.remove(url), log_textbox.delete("0.0", ctk.END), log_textbox.insert("0.0", "Delete complete."), frame.destroy()))
            delete_button.pack(side="left", padx=5, pady=5)

            product_id = iherb_url.rsplit('/', 1)[-1].replace("?rec=home", "")
            # product_id = re.search(r"/(\d+)$", iherb_url).group(1)
            product_id_button = ctk.CTkButton(url_frame, text=f"{product_id}", width=50)
            product_id_button.pack(side="left", pady=5)

            label = ctk.CTkLabel(url_frame, text=iherb_url)
            label.pack(side="left", padx=5, pady=5)

            return iherb_urls
# ================================================================================
def images_and_ingredients():
    amazon_iherb_value = amazon_iherb_option_var.get()

    if amazon_iherb_value == "amazon":

        from selenium.webdriver import ChromeOptions
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from webdriver_manager.chrome import ChromeDriverManager

        options = ChromeOptions()
        options.add_experimental_option("detach", True)
        # options.add_argument("--headless")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

        from selenium.webdriver.common.by import By
        import time
# =====

        '''

        # amazon sign-in 접속
        driver.get('https://www.amazon.com/-/ko/ap/signin?openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com%2F%3Flanguage%3Dko_KR%26ref_%3Dnav_ya_signin&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=usflex&openid.mode=checkid_setup&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&')

        # 전체화면으로 실행 (전체화면으로 실행해야 오류가 적음)
        driver.maximize_window()

        # e-mail, password 입력
        driver.find_element(By.CSS_SELECTOR, '#ap_email').send_keys('01099181244')
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, '#continue').click()

        driver.find_element(By.CSS_SELECTOR, '#ap_password').send_keys('30tp100djr1595')
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, '#signInSubmit').click()

        '''

        for url in amazon_urls:
            '''''''''''''''''''''''''''
            product image scraper
            '''''''''''''''''''''''''''
            # 해당 상품 주소로 이동
            time.sleep(1)
            driver.get(url)

            # 상품 이름 추출
            product_name = driver.find_element(By.CSS_SELECTOR, "#productTitle").text

            # ASIN CODE 추출
            import re 
# =====
            pattern = r'dp\/([A-Z0-9]{10})\/'
            asin_code = re.search(pattern, url).group(1)
            # print(asin_code)

            import os
# =====
            # 바탕화면에 있는 'kme' folder 안에 있는 'amazon' folder 안에 folder 만들기
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "amazon")
            folder_name = f"{asin_code}"
            folder_path = os.path.join(desktop_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            print(f"{asin_code} # folder created!")

            # 상품 사진 확대
            driver.find_element(By.CSS_SELECTOR, '#imgTagWrapperId').click()
            time.sleep(1)

            import requests
            from selenium.common.exceptions import NoSuchElementException
            from PIL import Image, ImageOps
# =====
            num_a = 0
            while True:
                css_selector = f"#ivImage_{num_a} > div"
                try:
                    # 사진 선택
                    driver.find_element(By.CSS_SELECTOR, css_selector).click()

                    # 선택한 사진 주소 추출
                    time.sleep(1)
                    image_url = driver.find_element(By.CSS_SELECTOR, '#ivLargeImage > img').get_attribute('src')

                    # 사진 원본 저장
                    response = requests.get(image_url)
                    filename = f"image{num_a + 1}.jpg"
                    print(image_url)
                    with open(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/{filename}", 'wb+') as f:
                        f.write(response.content)
                    print(f"{num_a + 1} # save complete!")

                    # 사진 열기
                    image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/{filename}")

                    if naver_checkbox.get():
                        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "amazon", f"{asin_code}")
                        folder_name = "naver(1000, 860)"
                        folder_path = os.path.join(desktop_path, folder_name)
                        os.makedirs(folder_path, exist_ok=True)
                        # print("# naver folder created!")

                        # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                        new_image = ImageOps.pad(image, (820, 820), color='white')

                        # 흰색 테두리 두께 설정 (위, 아래 20씩)
                        border_thickness = 20

                        # 테두리 두께 적용해서 사진 확장 (860x860)
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                        # 가공한 사진 저장
                        image_with_border.save(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/naver(1000, 860)/{filename}")
                        print(f"{num_a + 1} # retouch(naver) complete!")

                        '''''''''''''''''''''''''''
                        create thumbnail
                        '''''''''''''''''''''''''''
                        thumbnail_path_naver = f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/naver(1000, 860)/naver_thumbnail.jpg"

                        if not os.path.exists(thumbnail_path_naver):
                            # thumbnail 적용할 사진 불러오기
                            image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/naver(1000, 860)/image1.jpg")

                            # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                            new_image = ImageOps.pad(image, (960, 960), color='white')

                            # 흰색 테두리 두께 설정
                            border_thickness = 10

                            # 테두리 두께 적용해서 사진 확장하기
                            image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                            # 색상 테두리 두께 설정
                            border_thickness = 10

                            # 테두리 색상 설정
                            border_color = thumbnail_color
                            # print(border_color)

                            # 테두리 색상과 두께 적용해서 사진 확장하기
                            thumbnail_image_for_naver = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                            # thumbnail 저장하기
                            thumbnail_image_for_naver.save(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/naver(1000, 860)/naver_thumbnail.jpg")
                            print("= naver_thumbnail created!")
                        else:
                            pass

                    if coupang_checkbox.get():
                        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "amazon", f"{asin_code}")
                        folder_name = "coupang(500, 780)"
                        folder_path = os.path.join(desktop_path, folder_name)
                        os.makedirs(folder_path, exist_ok=True)
                        # print("# coupang folder created!")

                        # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                        new_image = ImageOps.pad(image, (740, 740), color='white')

                        # 흰색 테두리 두께 설정 (위, 아래 20씩)
                        border_thickness = 20

                        # 테두리 두께 적용해서 사진 확장 (780x780)
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                        # 가공한 사진 저장
                        image_with_border.save(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/coupang(500, 780)/{filename}")
                        print(f"{num_a + 1} # retouch(coupang) complete!")

                        '''''''''''''''''''''''''''
                        create thumbnail
                        '''''''''''''''''''''''''''
                        thumbnail_path_coupang = f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/coupang(500, 780)/coupang_thumbnail.jpg"

                        if not os.path.exists(thumbnail_path_coupang):
                            # thumbnail 적용할 사진 불러오기
                            image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/coupang(500, 780)/image1.jpg")

                            # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                            new_image = ImageOps.pad(image, (460, 460), color='white')

                            # 흰색 테두리 두께 설정
                            border_thickness = 10

                            # 테두리 두께 적용해서 사진 확장하기
                            image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                            # 색상 테두리 두께 설정
                            border_thickness = 10

                            # 테두리 색상 설정
                            border_color = thumbnail_color
                            # print(border_color)

                            # 테두리 색상과 두께 적용해서 사진 확장하기
                            thumbnail_image_for_coupang = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                            # thumbnail 저장하기
                            thumbnail_image_for_coupang.save(f"C:/Users/{os.getlogin()}/Desktop/kme/amazon/{asin_code}/coupang(500, 780)/coupang_thumbnail.jpg")
                            print("= coupang_thumbnail created!")
                        else:
                            pass

                    time.sleep(0.5)
                    
                    num_a = num_a + 1

                except NoSuchElementException:
                    print("----- no more image exist -----")
                    break

            time.sleep(0.5)
            for num_close in range(0, 10):
                css_selector = f"#a-popover-{num_close} > div > header > button"
                try:
                    driver.find_element(By.CSS_SELECTOR, css_selector).click()
                except NoSuchElementException:
                    pass
                    # print(f"element not found for {css_selector}")
                    
            log_textbox.delete("0.0", ctk.END)
            log_textbox.insert("0.0", "Image saved.")

            '''''''''''''''''''''''''''
            amazon OCR
            '''''''''''''''''''''''''''
            import pytesseract
# =====     
            warnings = []  # 누적해서 저장할 목록을 담을 변수

            for num_a in range(10):
                img_path = f"C:/Users/TILLIDIE/Desktop/kme/amazon/{asin_code}/image{num_a + 1}.jpg"
                
                if not os.path.exists(img_path):
                    # 사진이 존재하지 않으면 다음 순회로 건너뛰기
                    continue
                
                ocr_text = pytesseract.image_to_string(Image.open(img_path), lang='eng')
                
                with open('C:/Users/TILLIDIE/Desktop/kme/2022.10.16.txt', 'r') as f:
                    word_list = [line.strip() for line in f.readlines() if line.strip()]
                
                for word in word_list:
                    if word.lower() in ocr_text.lower():
                        ocr_words = ocr_text.split()
                        found = False
                        for ocr_word in ocr_words:
                            if word.lower() in ocr_word.lower():
                                ocr_word = ocr_word.replace("(", "").replace(")", "").replace(".", "").replace(",", "")
                                # warning_message = f"----- warning! [ {ocr_word} ] # {word} ----- from image{num_a + 1}"
                                warning_message = f"[ {ocr_word} ] # {word}"
                                warning_message = warning_message.replace("?", "").replace("_", "")
                                # 이전에 추가된 warning_message와 중복되는지 확인
                                if warning_message not in warnings:
                                    # 목록을 누적하여 저장
                                    warnings.append(warning_message)
                                found = True
                        if not found:
                            pass

            # 저장된 목록 출력 (금지성분이 발견되지 않은 사진은 표시되지 않음)
            print(f"\n{warnings}\n")

            # excel에 저장할 수 있도록 list를 문자열로 변환
            warnings_string = "\n".join(warnings)

            '''''''''''''''''''''''''''
            Insert row into Excel sheet
            '''''''''''''''''''''''''''
            path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook["amazon"]

            # 비어있는 행 식별 및 삭제
            empty_rows = []
            for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외
                if all(cell.value is None for cell in row):
                    empty_rows.append(row)

            for row in empty_rows:
                sheet.delete_rows(row[0].row)

            row_values = [asin_code, product_name, warnings_string, url]

            print(row_values)

            sheet.append(row_values)
            workbook.save(path)

            '''''''''''''''''''''''''''
            create product frame
            '''''''''''''''''''''''''''
            def display_image(image_path):
                image = Image.open(image_path)
                new_image = ImageOps.pad(image, (180, 180), color='white')
                border_thickness = 10
                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                photo = ImageTk.PhotoImage(image_with_border)
                product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                product_image_canvas.image = photo

            def warnings_string_textbox(image_path):
                asin_code = image_path.split("/")[-2]
                product_list = pd.read_excel("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx", sheet_name="amazon")
                filtered_row = product_list[product_list.iloc[:, 0] == asin_code]
                value = filtered_row.iloc[0, 2]
                
                import numpy as np

                if not isinstance(value, float) or not np.isnan(value):
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", value)
                else:
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", "")

            def warning_label(image_path):
                asin_code = image_path.split("/")[-2]

                print('asin_code_button_click ' + asin_code)

                product_list = pd.read_excel("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx", sheet_name="amazon")
                filtered_row = product_list[product_list.iloc[:, 0] == asin_code]
                value = filtered_row.iloc[0, 2]

                import numpy as np

                if not isinstance(value, float) or not np.isnan(value):
                    name_lable.configure(text="의심되는 성분이 존재합니다!")
                else:
                    name_lable.configure(text="금지성분이 발견되지 않았습니다.")

            def asin_code_button_click(image_path):
                # 제품 사진
                display_image(image_path)
                # 금지 성분
                warnings_string_textbox(image_path)
                # 제품 번호
                code_lable.configure(text=image_path.split("/")[-2])
                # 경고 문구
                warning_label(image_path)

            def pass_button_click(image_path, frame):
                asin_code = image_path.split("/")[-2]

                print('pass_button_click ' + asin_code)

                frame.configure(fg_color="#217346")

            def fail_button_click(image_path, frame):
                asin_code = image_path.split("/")[-2]

                print('fail_button_click ' + asin_code)

                log_textbox.delete("0.0", ctk.END)
                log_textbox.insert("0.0", "Delete complete.")
                frame.destroy()

                path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
                wb = openpyxl.load_workbook(path)
                sheet = wb["amazon"]

                for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                    if row[0].value == asin_code:
                        sheet.delete_rows(row[0].row)
                        break

                wb.save(path)

            # product frame
            product_frame = ctk.CTkFrame(scrollable_frame)
            product_frame.pack(fill="x", pady=(5,0))

            # asin code
            asin_code_button = ctk.CTkButton(product_frame, text=f"{asin_code}", width=50, 
                                             command=lambda img_path=f"C:/Users/TILLIDIE/Desktop/kme/amazon/{asin_code}/image1.jpg": 
                                             asin_code_button_click(img_path))
            asin_code_button.pack(side="left", padx=(5,0), pady=5)

            # pass button
            pass_button = ctk.CTkButton(product_frame, text="pass", width=50, fg_color="#217346", hover_color="#005000", 
                                        command=lambda img_path=f"C:/Users/TILLIDIE/Desktop/kme/amazon/{asin_code}/image1.jpg", 
                                        frame=product_frame: 
                                        pass_button_click(img_path, frame))
            pass_button.pack(side="left", padx=5, pady=5)

            # fail button
            fail_button = ctk.CTkButton(product_frame, text="fail", width=50, fg_color="#CC3D3D", hover_color="#960707", 
                                        command=lambda img_path=f"C:/Users/TILLIDIE/Desktop/kme/amazon/{asin_code}/image1.jpg", 
                                        frame=product_frame: 
                                        fail_button_click(img_path, frame))
            fail_button.pack(side="left", pady=5)

            # product name label
            label = ctk.CTkLabel(product_frame, text=product_name)
            label.pack(side="left", padx=5, pady=5, anchor="center")

        # web browser 종료
        driver.quit()

    else :

        import os
        import requests
        from bs4 import BeautifulSoup
# =====
        ''''''

        for url in iherb_urls:
            '''''''''''''''''''''''''''
            product image scraper
            '''''''''''''''''''''''''''
            # PRODUCT ID 추출
            import re
# =====     
            product_id = url.rsplit('/', 1)[-1].replace("?rec=home", "")
            # product_id = re.search(r"/(\d+)$", url).group(1)
            print(f"----- PRODUCT ID : {product_id} -----\n")

            # 바탕화면에 있는 'kme' folder 안에 있는 'iherb' folder 안에 folder 만들기
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "iherb")
            folder_name = f"{product_id}"
            folder_path = os.path.join(desktop_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            print(f"{product_id} # folder created!")

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
            }

            response = requests.get(url, headers=headers)
            soup = BeautifulSoup(response.content, "html.parser")

            img_tags = soup.find_all("img", {"data-large-img": True})

            from PIL import Image, ImageOps
# =====
            for num_i, img_tag in enumerate(img_tags):
                img_url = img_tag["data-large-img"]

                print(img_url)

                # 사진 이름 설정
                filename = f"image{num_i + 1}.jpg"

                # 사진 저장 요청 보내기
                img_response = requests.get(img_url, headers=headers)

                # 사진 원본 저장
                with open(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/{filename}", 'wb+') as f:
                        f.write(img_response.content)
                print(f"{num_i + 1} # save complete!")

                # 사진 열기
                image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/{filename}")

                if naver_checkbox.get():
                    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "iherb", f"{product_id}")
                    folder_name = "naver(1000, 860)"
                    folder_path = os.path.join(desktop_path, folder_name)
                    os.makedirs(folder_path, exist_ok=True)
                    # print("# naver folder created!")

                    # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                    new_image = ImageOps.pad(image, (820, 820), color='white')

                    # 흰색 테두리 두께 설정 (위, 아래 20씩)
                    border_thickness = 20

                    # 테두리 두께 적용해서 사진 확장 (860x860)
                    image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')           

                    # 가공한 사진 저장
                    image_with_border.save(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/naver(1000, 860)/{filename}")
                    print(f"{num_i + 1} # retouch(naver) complete!")

                    '''''''''''''''''''''''''''
                    create thumbnail
                    '''''''''''''''''''''''''''
                    thumbnail_path_naver = f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/naver(1000, 860)/naver_thumbnail.jpg"

                    if not os.path.exists(thumbnail_path_naver):
                        # thumbnail 적용할 사진 불러오기
                        image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/naver(1000, 860)/image1.jpg")

                        # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                        new_image = ImageOps.pad(image, (960, 960), color='white')

                        # 흰색 테두리 두께 설정
                        border_thickness = 10

                        # 테두리 두께 적용해서 사진 확장하기
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                        # 색상 테두리 두께 설정
                        border_thickness = 10

                        # 테두리 색상 설정
                        border_color = thumbnail_color
                        # print(border_color)

                        # 테두리 색상과 두께 적용해서 사진 확장하기
                        thumbnail_image_for_naver = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                        # thumbnail 저장하기
                        thumbnail_image_for_naver.save(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/naver(1000, 860)/naver_thumbnail.jpg")
                        print("= naver_thumbnail created!")
                    else:
                        pass

                if coupang_checkbox.get():
                    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "kme", "iherb", f"{product_id}")
                    folder_name = "coupang(500, 780)"
                    folder_path = os.path.join(desktop_path, folder_name)
                    os.makedirs(folder_path, exist_ok=True)
                    # print("# coupang folder created!")

                    # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                    new_image = ImageOps.pad(image, (740, 740), color='white')

                    # 흰색 테두리 두께 설정 (위, 아래 20씩)
                    border_thickness = 20

                    # 테두리 두께 적용해서 사진 확장 (780x780)
                    image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                    # 가공한 사진 저장
                    image_with_border.save(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/coupang(500, 780)/{filename}")
                    print(f"{num_i + 1} # retouch(coupang) complete!")

                    '''''''''''''''''''''''''''
                    create thumbnail
                    '''''''''''''''''''''''''''
                    thumbnail_path_coupang = f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/coupang(500, 780)/coupang_thumbnail.jpg"

                    if not os.path.exists(thumbnail_path_coupang):
                        # thumbnail 적용할 사진 불러오기
                        image = Image.open(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/coupang(500, 780)/image1.jpg")

                        # 사진 크기 조정 및 부족한 부분 흰색으로 채우기
                        new_image = ImageOps.pad(image, (460, 460), color='white')

                        # 흰색 테두리 두께 설정
                        border_thickness = 10

                        # 테두리 두께 적용해서 사진 확장하기
                        image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')

                        # 색상 테두리 두께 설정
                        border_thickness = 10

                        # 테두리 색상 설정
                        border_color = thumbnail_color
                        # print(border_color)

                        # 테두리 색상과 두께 적용해서 사진 확장하기
                        thumbnail_image_for_coupang = ImageOps.expand(image_with_border, border=border_thickness, fill=border_color)

                        # thumbnail 저장하기
                        thumbnail_image_for_coupang.save(f"C:/Users/{os.getlogin()}/Desktop/kme/iherb/{product_id}/coupang(500, 780)/coupang_thumbnail.jpg")
                        print("= coupang_thumbnail created!")
                    else:
                        pass
            
            log_textbox.delete("0.0", ctk.END)
            log_textbox.insert("0.0", "Image saved.")

            # 상품 이름 추출
            product_name = soup.find('div', class_='product-summary-title').get_text(strip=True)
            print(product_name)

            '''''''''''''''''''''''''''
            iherb OCR
            '''''''''''''''''''''''''''
            warnings = []  # 누적해서 저장할 목록을 담을 변수

            # 영양 성분 정보
            container = soup.find('div', {'class': 'supplement-facts-container'})
            supplement_facts_text = container.get_text(strip=True)
            supplement_facts_text = supplement_facts_text.replace('영양 성분 정보', '영양 성분 정보\n')
            # print("\n ----- 영양 성분 정보 -----")
            # print(supplement_facts_text)

            '''

            # 상품 사용법
            container = soup.find('div', {'class': 'prodOverviewDetail'})
            suggested_use_text = container.get_text(strip=True)
            print("----- 상품 사용법 -----")
            print(suggested_use_text)

            '''

            # 포함된 다른 성분들
            container = soup.find('div', {'class': 'prodOverviewIngred'})
            other_ingredients_text = container.get_text(strip=True)
            other_ingredients_text = other_ingredients_text.replace('주요 성분', '주요 성분\n')
            other_ingredients_text = other_ingredients_text.replace('기타 성분', '\n기타 성분\n')
            if '무함유.' in other_ingredients_text:
                other_ingredients_text = '무함유.'
                # print("----- 포함된 다른 성분들 -----")
                # print(other_ingredients_text)
            elif '이 제품은' in other_ingredients_text:
                other_ingredients_text = other_ingredients_text.split('이 제품은')[0]
                # print("----- 포함된 다른 성분들 -----")
                # print(other_ingredients_text)
            elif '이 제품에는' in other_ingredients_text:
                other_ingredients_text = other_ingredients_text.split('이 제품에는')[0]
                # print("----- 포함된 다른 성분들 -----")
                # print(other_ingredients_text)
            else:
                pass
                # print("----- 포함된 다른 성분들 -----")
                # print(other_ingredients_text)
                
            from googletrans import Translator
# =====
            translator = Translator()
            import time
            time.sleep(0.5)

            # supplement_facts_text에서 '%하루 영양소 기준치' 이후의 문자열 추출
            match = re.search(r'%하루 영양소 기준치(.*?)$', supplement_facts_text, re.DOTALL)
            if match:
                extracted_text = match.group(1).strip()
            supplement_facts_text = extracted_text

            # 번역할 문자열
            text_to_translate = supplement_facts_text + " " + other_ingredients_text

            # 영어로 번역
            translated_text = translator.translate(text_to_translate, dest='en').text
            translated_text = translated_text.replace("*", "").replace("The standard value per day is not set.", ", ")
            translated_text = translated_text.replace("Nothing.", "")

            # 번역문을 출력
            # print(f"\n{translated_text}\n")

            # 검사할 단어 목록을 메모장에서 불러오기
            with open('C:/Users/TILLIDIE/Desktop/kme/2022.10.16.txt', 'r') as f:
                word_list = [line.strip() for line in f.readlines() if line.strip()]

            # 번역문을 검사할 단어 목록과 비교하여 겹치는 단어가 있으면 출력
            found = False
            for word in word_list:
                if word.lower() in translated_text.lower():
                    ocr_words = translated_text.split()
                    for ocr_word in ocr_words:
                        if word.lower() in ocr_word.lower():
                            ocr_word = ocr_word.replace("(", "").replace(")", "").replace(".", "").replace(",", "")
                            warning_message = f"[ {ocr_word} ] # {word}"
                            warning_message = warning_message.replace("?", "").replace("_", "")
                            # 이전에 추가된 warning_message와 중복되는지 확인
                            if warning_message not in warnings:
                                # 목록을 누적하여 저장
                                warnings.append(warning_message)
                            found = True
                    if not found:
                        pass

            # 저장된 목록 출력 (금지성분이 발견되지 않은 사진은 표시되지 않음)
            print(f"\n{warnings}\n")

            # excel에 저장할 수 있도록 list를 문자열로 변환
            warnings_string = "\n".join(warnings)

            '''''''''''''''''''''''''''
            Insert row into Excel sheet
            '''''''''''''''''''''''''''
            path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook["iherb"]

            # 비어있는 행 식별 및 삭제
            empty_rows = []
            for row in sheet.iter_rows(min_row=2):  # 첫 번째 행은 제목이므로 제외
                if all(cell.value is None for cell in row):
                    empty_rows.append(row)

            for row in empty_rows:
                sheet.delete_rows(row[0].row)

            row_values = [product_id, product_name, warnings_string, url]

            print(row_values)

            sheet.append(row_values)
            workbook.save(path)

            '''''''''''''''''''''''''''
            create product frame
            '''''''''''''''''''''''''''
            def iherb_display_image(image_path):
                image = Image.open(image_path)
                new_image = ImageOps.pad(image, (180, 180), color='white')
                border_thickness = 10
                image_with_border = ImageOps.expand(new_image, border=border_thickness, fill='white')
                photo = ImageTk.PhotoImage(image_with_border)
                product_image_canvas.create_image(100, 100, anchor="center", image=photo)
                product_image_canvas.image = photo

            def iherb_warnings_string_textbox(image_path):
                product_id = image_path.split("/")[-2]
                iherb_product_list = pd.read_excel("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx", sheet_name="iherb")
                iherb_filtered_row = iherb_product_list[iherb_product_list.iloc[:, 0].astype(str) == str(product_id)]
                iherb_value = iherb_filtered_row.iloc[0, 2]
                
                import numpy as np

                if not isinstance(iherb_value, float) or not np.isnan(iherb_value):
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", iherb_value)
                else:
                    ingredients_textbox.delete("0.0", ctk.END)
                    ingredients_textbox.insert("0.0", "")

            def iherb_warning_label(image_path):
                product_id = image_path.split("/")[-2]

                print('product_id_button_click ' + product_id)

                iherb_product_list = pd.read_excel("C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx", sheet_name="iherb")
                iherb_filtered_row = iherb_product_list[iherb_product_list.iloc[:, 0].astype(str) == str(product_id)]
                iherb_value = iherb_filtered_row.iloc[0, 2]

                import numpy as np

                if not isinstance(iherb_value, float) or not np.isnan(iherb_value):
                    name_lable.configure(text="의심되는 성분이 존재합니다!")
                else:
                    name_lable.configure(text="금지성분이 발견되지 않았습니다.")

            def product_id_button_click(image_path):
                # 제품 사진
                iherb_display_image(image_path)
                # 금지 성분
                iherb_warnings_string_textbox(image_path)
                # 제품 번호
                code_lable.configure(text=image_path.split("/")[-2])
                # 경고 문구
                iherb_warning_label(image_path)

            def iherb_pass_button_click(image_path, frame):
                product_id = image_path.split("/")[-2]

                print('iherb_pass_button_click ' + product_id)

                frame.configure(fg_color="#217346")

            def iherb_fail_button_click(image_path, frame):
                product_id = image_path.split("/")[-2]

                print('iherb_fail_button_click ' + product_id)

                log_textbox.delete("0.0", ctk.END)
                log_textbox.insert("0.0", "Delete complete.")
                frame.destroy()

                path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
                wb = openpyxl.load_workbook(path)
                sheet = wb["iherb"]

                for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                    if row[0].value == product_id:
                        sheet.delete_rows(row[0].row)
                        break

                wb.save(path)

            # product frame
            product_frame = ctk.CTkFrame(scrollable_frame)
            product_frame.pack(fill="x", pady=(5,0))

            # product id
            product_id_button = ctk.CTkButton(product_frame, text=f"{product_id}", width=50,
                                              command=lambda img_path=f"C:/Users/TILLIDIE/Desktop/kme/iherb/{product_id}/image1.jpg":
                                              product_id_button_click(img_path))
            product_id_button.pack(side="left", padx=(5,0), pady=5)

            # pass button
            pass_button = ctk.CTkButton(product_frame, text="pass", width=50, fg_color="#217346", hover_color="#005000", 
                                        command=lambda img_path = f"C:/Users/TILLIDIE/Desktop/kme/iherb/{product_id}/image1.jpg", 
                                        frame=product_frame: 
                                        iherb_pass_button_click(img_path, frame))
            pass_button.pack(side="left", padx=5, pady=5)

            # fail button
            fail_button = ctk.CTkButton(product_frame, text="fail", width=50, fg_color="#CC3D3D", hover_color="#960707", 
                                        command=lambda img_path = f"C:/Users/TILLIDIE/Desktop/kme/iherb/{product_id}/image1.jpg", 
                                        frame=product_frame: 
                                        iherb_fail_button_click(img_path, frame))
            fail_button.pack(side="left", pady=5)

            # product name label
            label = ctk.CTkLabel(product_frame, text=product_name, font=font_style)
            label.pack(side="left", padx=5, pady=5, anchor="center")
# ================================================================================
def save_to_a_database():
    amazon_iherb_value = amazon_iherb_option_var.get()

    if amazon_iherb_value == "amazon":
        path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["amazon"]

        # 비어있는 행 식별 및 삭제
        empty_rows = []
        for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외
            if all(cell.value is None for cell in row):
                empty_rows.append(row)

        for row in empty_rows:
            sheet.delete_rows(row[0].row)

        workbook.save(path)

    else:
        path = "C:/Users/TILLIDIE/Desktop/kme/product_list.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["iherb"]

        # 비어있는 행 식별 및 삭제
        empty_rows = []
        for row in sheet.iter_rows(min_row=2): # 첫 번째 행은 제목이므로 제외
            if all(cell.value is None for cell in row):
                empty_rows.append(row)

        for row in empty_rows:
            sheet.delete_rows(row[0].row)

        workbook.save(path)
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
# ui
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
# ================================================================================
root = ctk.CTk()
root.title("KME Scraper")
# ================================================================================

ctk.set_appearance_mode("dark") # system, dark, light
ctk.set_default_color_theme("blue")  # blue(standard), green, dark-blue

font_style = ctk.CTkFont("돋움")

''''''

# style = ttk.Style(root)
# root.tk.call("source", 
#              "C:/Users/TILLIDIE/Desktop/CODING/VS Code/tkinter/customtkinter_02/forest-light.tcl")
# root.tk.call("source", 
#              "C:/Users/TILLIDIE/Desktop/CODING/VS Code/tkinter/customtkinter_02/forest-dark.tcl")
# style.theme_use("forest-dark")

''''''

# ================================================================================
frame_01 = ctk.CTkFrame(root)
frame_01.grid(row=0, column=0, padx=(20,10), pady=(20,10), sticky="news")
# ================================================================================

option_frame = ctk.CTkFrame(frame_01)
option_frame.pack(side="left", fill="both", expand=True, padx=(20,10), pady=20)

# amazon or iherb radiobutton
amazon_iherb_option_var = ctk.StringVar(value="amazon")
amazon_radiobutton = ctk.CTkRadioButton(option_frame, text="amazon", 
                                        variable=amazon_iherb_option_var, value="amazon")
amazon_radiobutton.pack(side="top", fill="both", expand=True, padx=20, pady=(20,5))
iherb_radiobutton = ctk.CTkRadioButton(option_frame, text="iherb", 
                                       variable=amazon_iherb_option_var, value="iherb")
iherb_radiobutton.pack(side="bottom", fill="both", expand=True, padx=20, pady=(5,20))

# amazon_iherb_option = ctk.CTkOptionMenu(option_frame, values=["none", "amazon", "iherb"])
# amazon_iherb_option.set("none")
# amazon_iherb_option.pack(expand=True, padx=20, pady=20)

''''''

entry_frame = ctk.CTkFrame(frame_01)
entry_frame.pack(fill="both", side="right", expand=True, padx=(10,20), pady=20)

# amazon id entry
id_entry = ctk.CTkEntry(entry_frame, placeholder_text="identification") # text box to input
id_entry.pack(fill="x", expand=True, side="top", padx=20, pady=(20,5))

# amazon password entry
pw_entry = ctk.CTkEntry(entry_frame, placeholder_text="password") # text box to input
pw_entry.pack(fill="x", expand=True, side="bottom", padx=20, pady=(5,20))

# ================================================================================
frame_02 = ctk.CTkFrame(root)
frame_02.grid(row=1, column=0, padx=(20,10), pady=10, sticky="news")
# ================================================================================

image_frame = ctk.CTkFrame(frame_02)
image_frame.pack(side="top", fill="both", expand=True, padx=20, pady=(20,10))

# naver checkbox
naver_checkbox = ctk.CTkCheckBox(image_frame, text="naver (1000, 860)")
naver_checkbox.pack(expand=True, side="left", padx=(50,10), pady=10)
# coupang checkbox
coupang_checkbox = ctk.CTkCheckBox(image_frame, text="coupang (500, 780)")
coupang_checkbox.pack(expand=True, side="left", padx=(10,50), pady=10)

''''''

border_frame = ctk.CTkFrame(frame_02)
border_frame.pack(side="top", fill="both", expand=True, padx=20, pady=(0,20))

# thumbnail color segemented button
color_segemented_button = ctk.CTkSegmentedButton(border_frame, 
                                                 values=[" white ", " red ", " orange ", " yellow ", " green ", " blue ", " purple "], 
                                                 command=color_segmented_button_callback)
color_segemented_button.set(" white ")
color_segemented_button.pack(expand=True, padx=10, pady=10)

# ================================================================================
frame_03 = ctk.CTkFrame(root)
frame_03.grid(row=0, rowspan=3, column=1, padx=(10,20), pady=(20,10), sticky="news")
# ================================================================================
url_entry = ctk.CTkEntry(frame_03, placeholder_text="product URL")
url_entry.pack(fill="x", padx=20, pady=(20,5))

url_scrollable_frame = ctk.CTkScrollableFrame(frame_03, width=500)
url_scrollable_frame.pack(fill="y", expand=True, padx=20, pady=(5,10))

# check for duplicates and add product URL button
add_button = ctk.CTkButton(frame_03, width=200, text="check for duplicates and add product URL", 
                           command=check_duplicates_and_add_URL)
add_button.pack(fill="x", padx=20, pady=(0,5))

# collect product images and inspect ingredients button
add_button = ctk.CTkButton(frame_03, width=200, text="collect product images and inspect ingredients", 
                           command=images_and_ingredients)
add_button.pack(fill="x", padx=20, pady=(5,20))

# ================================================================================
textbox_frame = ctk.CTkFrame(root)
textbox_frame.grid(row=2, column=0, padx=(20,10), pady=10, sticky="ews")
# ================================================================================

# log textbox
log_textbox = ctk.CTkTextbox(textbox_frame, height=30, font=ctk.CTkFont("돋움"))
log_textbox.pack(fill="both", expand=True, padx=20, pady=10)

# ================================================================================
frame_04 = ctk.CTkFrame(root)
frame_04.grid(row=3, column=0, padx=(20,10), pady=(10,10), sticky="news")
# ================================================================================

# product image canvas
product_image_canvas = tk.Canvas(frame_04, width=200, height=200)
product_image_canvas.grid(row=0, column=0, padx=(20, 10), pady=(20,10))

''''''

ingredients_textbox_frame = ctk.CTkFrame(frame_04)
ingredients_textbox_frame.grid(row=0, column=1, padx=(10,20), pady=(20,10), sticky="news")

# ingredients textbox
ingredients_textbox = ctk.CTkTextbox(ingredients_textbox_frame, activate_scrollbars=False, font=font_style)
ingredients_textbox.pack(side="left", expand=True, fill="both", padx=(5,0), pady=5)

ctk_textbox_scrollbar = ctk.CTkScrollbar(ingredients_textbox_frame, command=ingredients_textbox.yview)
ctk_textbox_scrollbar.pack(side="right", fill="y")

ingredients_textbox.configure(yscrollcommand=ctk_textbox_scrollbar.set)

''''''

product_code = ctk.CTkFrame(frame_04, height=50)
product_code.grid(row=1, column=0, padx=(20,10), pady=(0,20), sticky="ew")

# code label
code_lable = ctk.CTkLabel(product_code, text="")
code_lable.pack(padx=10, pady=10)

''''''

pass_fail_frame = ctk.CTkFrame(frame_04, height=50)
pass_fail_frame.grid(row=1, column=1, padx=(10,20), pady=(0,20), sticky="ew")

name_lable = ctk.CTkLabel(pass_fail_frame, text="", width=250, font=font_style)
name_lable.pack(padx=10, pady=10)

# # pass button
# pass_button = ctk.CTkButton(pass_fail_frame, text="pass", 
#                             fg_color="#217346", hover_color="#005000")
# pass_button.pack(side="left", fill="x", padx=(10,5), pady=10)

# # fail button
# fail_button = ctk.CTkButton(pass_fail_frame, text="fail", 
#                             fg_color="#CC3D3D", hover_color="#960707")
# fail_button.pack(side="right", fill="x", padx=(5,10), pady=10)

# ================================================================================
frame_05 = ctk.CTkFrame(root)
frame_05.grid(row=4, column=0, padx=(20,10), pady=10, sticky="news")
# ================================================================================

license_frame = ctk.CTkFrame(frame_05)
license_frame.pack(side="left", fill="x", expand=True, padx=(20,5), pady=10)

# license label
license_lable = ctk.CTkLabel(license_frame, text="DESKTOP-LNMR5Q3")
license_lable.pack(padx=5)

''''''

list_date_frame = ctk.CTkFrame(frame_05)
list_date_frame.pack(side="left", fill="x", expand=True, padx=5, pady=10)

# list date label
list_date_lable = ctk.CTkLabel(list_date_frame, text="list date. 2022.10.16")
list_date_lable.pack(padx=5)

''''''

version_frame = ctk.CTkFrame(frame_05)
version_frame.pack(side="left", fill="x", expand=True, padx=(5,20), pady=10)

# version label
version_lable = ctk.CTkLabel(version_frame, text="version. 23.1.0")
version_lable.pack(padx=5)

# ================================================================================
frame_06 = ctk.CTkFrame(root)
frame_06.grid(row=3, rowspan=2, column=1, padx=(10,20), pady=10, sticky="news")
# ================================================================================

scrollable_frame = ctk.CTkScrollableFrame(frame_06, width=500)
scrollable_frame.pack(fill="y", expand=True, padx=20, pady=(20,10))

# save to a database button
save_button = ctk.CTkButton(frame_06, width=200, text="save to a database", 
                            command=save_to_a_database)
save_button.pack(fill="x", padx=20, pady=(0,20))

# ================================================================================
frame_07 = ctk.CTkFrame(root)
frame_07.grid(row=5, column=0, columnspan=2, padx=20, pady=(10,20), sticky="news")
# ================================================================================

code_entry = ctk.CTkEntry(frame_07, placeholder_text="product code") # text box to input
code_entry.pack(fill="x", expand=True, side="left", padx=(20,5), pady=10)

# search button
search_button = ctk.CTkButton(frame_07, text="search")
search_button.pack(fill="x", expand=True, side="left", padx=5, pady=10)

# url textbox
url_textbox = ctk.CTkTextbox(frame_07, width=600, height=30)
url_textbox.pack(fill="x", expand=True, side="left", padx=5, pady=10)

# copy URL button
copy_url_button = ctk.CTkButton(frame_07, text="copy URL")
copy_url_button.pack(fill="x", expand=True, side="left", padx=(5,20), pady=10)

''''''

root.mainloop()
